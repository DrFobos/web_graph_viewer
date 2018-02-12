import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
from openpyxl import load_workbook


# read excel file and create two lists of nodes and cleared edges
def read_excel_file(filename, clear_branches_func):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb[wb.sheetnames[0]]

    nodes = []
    branches = []
    current_row = 0

    for row in ws.rows:
        current_row += 1
        if current_row >= 4:
            if not (row[2].value is None or row[2].value in nodes):
                nodes.append(row[2].value)
            if not row[0].value is None and not {row[0].value, row[1].value} in branches:
                branches.append({row[0].value, row[1].value})

    return nodes, clear_branches_func(nodes, branches)


# remove bad edges from list of them return list of edges
def clear_branches(nodes, edges):
    for edge in edges:
        if len(edge) != 2:
            edges.remove(edge)
        else:
            for node in edge:
                if node not in nodes:
                    edges.remove(edge)
                    break

    return list(map(lambda x: list(x), edges))


# create adjacency matrix from list of nodes and edges
# return adjacency matrix and dictionary of labels
def adjacency_matrix_generator(nodes, branches):
    id_dict = {}

    nodes_number = len(nodes)
    nodes.sort()

    for i in range(nodes_number):
        id_dict[nodes[i]] = i

    adjacency_matrix = np.zeros((nodes_number, nodes_number), dtype=int)

    for branch in branches:
        i = id_dict[branch[0]]
        j = id_dict[branch[1]]
        adjacency_matrix[i][j] = 1
        adjacency_matrix[j][i] = 1

    labels = {v: str(k) for k, v in id_dict.items()}

    return adjacency_matrix, labels


# draw graph from adjacency matrix using passed labels
def show_graph_with_labels(adjacency_matrix, mylabels=None):
    rows, cols = np.where(adjacency_matrix == 1)
    edges = zip(rows.tolist(), cols.tolist())
    gr = nx.Graph()
    gr.add_edges_from(edges)

    nx.draw(gr, node_size=1600, labels=mylabels, with_labels=True, node_color='deepskyblue', edge_color='darkgreen',
            pos=nx.shell_layout(gr), node_alpha=0.5, node_text_size=16,
            edge_alpha=0.3, edge_tickness=50)

    plt.savefig("Graph.png", format="PNG")
    plt.show()


if __name__ == '__main__':
    nodes, branches = read_excel_file('Data.xlsx', clear_branches)
    am, id_dict = adjacency_matrix_generator(nodes, branches)
    show_graph_with_labels(am, id_dict)
